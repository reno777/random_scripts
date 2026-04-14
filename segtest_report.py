#!/usr/bin/env python3
"""
segtest_report.py — Segmentation Test Report Generator
=======================================================

PURPOSE
-------
Converts raw nmap XML scan output (and optional EyeWitness SQLite databases)
into a formatted Excel workbook suitable for PCI/segmentation test reporting.
Results are organized by scanning location (perspective) so that findings from
different network segments can be reviewed and compared side-by-side.

USAGE
-----
    python segtest_report.py [directory] [-o output.xlsx]

    directory   Path to the folder containing your scan files. The script
                searches recursively, so nested subdirectories are fine.
                Defaults to the current directory if omitted.

    -o FILE     Output file name or full path. If only a filename is given
                (e.g. -o report.xlsx) the file is written into the scanned
                directory. Defaults to segmentation_report.xlsx in the
                scanned directory.

EXAMPLES
--------
    # Search current directory, write report here
    python segtest_report.py

    # Search a specific results folder
    python segtest_report.py /engagements/acme/scans

    # Custom output path
    python segtest_report.py /engagements/acme/scans -o /engagements/acme/report.xlsx

NMAP REQUIREMENTS
-----------------
    Any nmap output format works as a source, but the script parses XML only.
    If you used -oA (recommended), the .xml file is ready to use as-is.
    Service version info (-sV) and OS detection (-O) populate additional
    columns when present. NSE script output (-sC or --script) is captured
    in the "NSE Script Output" column.

EYEWITNESS INTEGRATION
----------------------
    Run EyeWitness against your nmap XML before running this script:
        eyewitness -x scan.xml -d ew_results/

    Point this script at the parent directory — it will find the ewdb.db or
    eyewitness.db file automatically and join web titles / HTTP status codes
    onto matching IP+port rows.

WORKFLOW (interactive)
----------------------
    1. Script discovers all nmap XML files and EyeWitness DBs under the
       given directory.
    2. You are prompted to assign each file to a named scanning location
       (e.g. Azure, Mason, Lansing, Detroit).
         • Multiple files assigned to the same location are merged and
           deduplicated — useful if you ran multiple scans from one segment.
         • Press Enter to skip a file you don't want to include.
    3. An Excel workbook is produced with:
         • "All Results"  — every finding across all locations
         • One tab per location (e.g. "Azure", "Mason", ...)
         • "Legend"       — colour key and flagged port reference lists

COLOUR CODING
-------------
    Red    — Sensitive / high-risk port, state: open
    Orange — Sensitive / high-risk port, state: open|filtered
    Blue   — Web / HTTP service port (any state)
    (none) — Standard port; alternating rows shaded light grey for readability

DEPENDENCIES
------------
    pip install openpyxl
"""

import argparse
import ipaddress
import re
import sqlite3
import sys
import xml.etree.ElementTree as ET
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit(
        "ERROR: openpyxl is required.\n"
        "       Install with:  pip install openpyxl"
    )

# ---------------------------------------------------------------------------
# Port classification
# ---------------------------------------------------------------------------
#
# Ports are classified as sensitive based on protocol so that, for example,
# TCP 512-514 (rexec/rlogin/rsh) flag correctly while UDP 512-514 do not.
# Add or remove entries here to match the scope of your engagement.

# TCP ports commonly flagged in PCI/segmentation reviews
SENSITIVE_PORTS_TCP = {
    21,    # FTP
    22,    # SSH
    23,    # Telnet
    25,    # SMTP
    53,    # DNS (zone transfer risk)
    110,   # POP3
    111,   # RPC portmapper
    135,   # MS-RPC endpoint mapper
    139,   # NetBIOS session service
    143,   # IMAP
    389,   # LDAP
    445,   # SMB
    512,   # rexec
    513,   # rlogin
    514,   # rsh / syslog (TCP)
    873,   # rsync
    1433,  # MS SQL Server
    1521,  # Oracle DB
    2049,  # NFS
    2375,  # Docker daemon (unauthenticated)
    2376,  # Docker daemon (TLS)
    3306,  # MySQL / MariaDB
    3389,  # RDP
    4444,  # Common reverse-shell / Metasploit default
    5432,  # PostgreSQL
    5900,  # VNC
    5985,  # WinRM (HTTP)
    5986,  # WinRM (HTTPS)
    6379,  # Redis (unauthenticated by default)
    9200,  # Elasticsearch REST API
    11211, # Memcached
    27017, # MongoDB
    27018, # MongoDB (shard)
}

# UDP ports commonly flagged in PCI/segmentation reviews
SENSITIVE_PORTS_UDP = {
    53,   # DNS
    69,   # TFTP (no authentication)
    111,  # RPC portmapper
    137,  # NetBIOS name service
    138,  # NetBIOS datagram service
    161,  # SNMP (community strings)
    162,  # SNMP trap
    500,  # IKE / IPsec
    1900, # SSDP / UPnP
    2049, # NFS
    4500, # IPsec NAT traversal
}

# Web-facing ports — lower severity than sensitive ports but worth calling out
WEB_PORTS = {80, 443, 8000, 8008, 8080, 8088, 8443, 8888, 9000, 9090, 9443}

# ---------------------------------------------------------------------------
# Colour palette (openpyxl PatternFill)
# ---------------------------------------------------------------------------

RED_FILL    = PatternFill("solid", fgColor="FFB3B3")  # sensitive port, open
ORANGE_FILL = PatternFill("solid", fgColor="FFD9A0")  # sensitive port, open|filtered
BLUE_FILL   = PatternFill("solid", fgColor="BDD7EE")  # web port
ALT_FILL    = PatternFill("solid", fgColor="DCE6F1")  # alternating row shading

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")  # dark blue header background
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)

# Thin grey border applied to every cell for clean table rendering
_thin  = Side(style="thin", color="AAAAAA")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def port_fill(port_str: str, protocol: str, state: str):
    """
    Return the appropriate PatternFill for a given port/protocol/state
    combination, or None if no highlight should be applied.

    Priority order: sensitive (red/orange) > web (blue) > no fill.
    """
    try:
        p = int(port_str)
    except (ValueError, TypeError):
        return None

    proto       = protocol.lower()
    is_open     = state == "open"
    is_filtered = state == "open|filtered"

    # Check sensitivity using the correct protocol-specific set
    sensitive = (
        (proto == "tcp" and p in SENSITIVE_PORTS_TCP) or
        (proto == "udp" and p in SENSITIVE_PORTS_UDP)
    )

    if sensitive and is_open:
        return RED_FILL
    if sensitive and is_filtered:
        return ORANGE_FILL
    if p in WEB_PORTS:
        return BLUE_FILL
    return None


# ---------------------------------------------------------------------------
# Scope file parser
# ---------------------------------------------------------------------------

def _count_token(token: str) -> int:
    """
    Return the number of IP targets represented by a single scope token.

    Handles the following formats (same as nmap -iL):
      192.168.1.1           — single IP → 1
      192.168.1.0/24        — CIDR      → num_addresses (256 for /24)
      192.168.1.1-50        — short dash range (last-octet end) → 50
      192.168.1.1-192.168.1.50 — full dash range → 50
      db.internal           — hostname  → 1 (counted as-is)
    """
    token = token.strip()
    if not token:
        return 0

    # ── CIDR notation ────────────────────────────────────────────────────────
    try:
        net = ipaddress.ip_network(token, strict=False)
        return net.num_addresses
    except ValueError:
        pass

    # ── Dash range ───────────────────────────────────────────────────────────
    if "-" in token:
        parts = token.split("-", 1)
        try:
            start_ip = ipaddress.ip_address(parts[0])
            # Try a fully-qualified end IP first (e.g. 10.0.0.1-10.0.0.50)
            try:
                end_ip = ipaddress.ip_address(parts[1])
                count  = int(end_ip) - int(start_ip) + 1
                return max(count, 1)
            except ValueError:
                pass
            # Fall back to short-form end (e.g. 192.168.1.1-50):
            # replace the last octet of the start IP with the end value
            start_octets = parts[0].split(".")
            if len(start_octets) == 4 and parts[1].isdigit():
                end_ip_str = ".".join(start_octets[:-1]) + "." + parts[1]
                try:
                    end_ip = ipaddress.ip_address(end_ip_str)
                    count  = int(end_ip) - int(start_ip) + 1
                    return max(count, 1)
                except ValueError:
                    pass
        except ValueError:
            pass

    # ── Hostname or unrecognised format — count as one target ────────────────
    return 1


def parse_scope_file(scope_path: Path) -> int:
    """
    Parse a scope file and return the total number of targets it defines.

    The file format follows nmap -iL conventions:
      • One or more targets per line (whitespace-separated)
      • Lines beginning with # are treated as comments and skipped
      • Blank lines are skipped
      • Supported per-token formats: single IP, CIDR, dash range, hostname

    Returns the total target count as an integer.
    """
    total = 0
    try:
        with open(scope_path, "r") as fh:
            for line in fh:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                # nmap -iL allows multiple space-separated targets on one line
                for token in line.split():
                    total += _count_token(token)
    except OSError as exc:
        print(f"  WARNING: could not read scope file ({scope_path}): {exc}")
    return total


def prompt_scope_file() -> int | None:
    """
    Interactively ask the user for a scope file path.

    Returns the parsed target count, or None if the user skips.
    The prompt retries if the given path does not exist.
    """
    print(f"\n{'=' * 64}")
    print("  Scope file (used for accurate 'IPs Targeted' count).")
    print("  The file should list targets in nmap -iL format")
    print("  (IPs, CIDRs, dash ranges, or hostnames — one or more per line).")
    print("  Press Enter to skip (will fall back to nmap runstats).\n")

    while True:
        raw = input("  Path to scope file: ").strip()
        if not raw:
            print("  Skipping scope file — IPs Targeted will use nmap runstats.\n")
            return None

        # Expand ~ and resolve the path
        scope_path = Path(raw).expanduser().resolve()
        if not scope_path.exists():
            print(f"  File not found: {scope_path}  (try again or press Enter to skip)\n")
            continue

        count = parse_scope_file(scope_path)
        print(f"  Scope file parsed: {count:,} target(s) found in {scope_path.name}\n")
        return count


# ---------------------------------------------------------------------------
# File discovery
# ---------------------------------------------------------------------------

def is_nmap_xml(path: Path) -> bool:
    """Return True if the file is a valid nmap XML output file."""
    try:
        root = ET.parse(path).getroot()
        return root.tag == "nmaprun"
    except Exception:
        return False


def find_nmap_xmls(directory: Path) -> list[Path]:
    """Recursively find all nmap XML files under directory."""
    return sorted(p for p in directory.rglob("*.xml") if is_nmap_xml(p))


def find_eyewitness_dbs(directory: Path) -> list[Path]:
    """
    Recursively find EyeWitness SQLite databases.
    EyeWitness names the DB either ewdb.db or eyewitness.db depending on
    the version, so both filenames are checked.
    """
    found = []
    for name in ("ewdb.db", "eyewitness.db"):
        found.extend(directory.rglob(name))
    return sorted(set(found))


# ---------------------------------------------------------------------------
# nmap XML parser
# ---------------------------------------------------------------------------

def parse_nmap_xml(xml_path: Path) -> list[dict]:
    """
    Parse a single nmap XML file and return a list of port-level records.

    One record is produced per open/open|filtered port per host. If a host
    is up but has no qualifying ports, a single placeholder record is still
    created so the host appears in the report (state: "up (no open ports)").
    """
    records: list[dict] = []
    try:
        root = ET.parse(xml_path).getroot()
    except ET.ParseError as exc:
        print(f"  WARNING: cannot parse {xml_path.name}: {exc}")
        return records

    # Store the original scan arguments for provenance (available in nmap XML)
    scan_args = root.get("args", "")

    for host in root.findall("host"):
        # Skip hosts that did not respond
        status_el = host.find("status")
        if status_el is None or status_el.get("state") != "up":
            continue

        ip = mac = hostname = os_name = ""

        # Pull IP (v4 or v6) and MAC from the address elements
        for addr in host.findall("address"):
            atype = addr.get("addrtype", "")
            if atype in ("ipv4", "ipv6"):
                ip = addr.get("addr", "")
            elif atype == "mac":
                mac = addr.get("addr", "")

        if not ip:
            continue

        # Use the first hostname nmap resolved (PTR or user-supplied)
        hn_el = host.find("hostnames")
        if hn_el is not None:
            first = hn_el.find("hostname")
            if first is not None:
                hostname = first.get("name", "")

        # OS detection — use the top-ranked osmatch if present
        os_el = host.find("os")
        if os_el is not None:
            om = os_el.find("osmatch")
            if om is not None:
                os_name = om.get("name", "")

        ports_el = host.find("ports")
        if ports_el is None:
            # Host is up but nmap returned no ports element at all
            records.append(_base_record(ip, hostname, mac, os_name, scan_args))
            continue

        open_found = False
        for port in ports_el.findall("port"):
            state_el = port.find("state")
            state = state_el.get("state", "") if state_el is not None else ""

            # Only include ports that are open or ambiguous (open|filtered)
            if state not in ("open", "open|filtered"):
                continue

            open_found = True
            port_id  = port.get("portid", "")
            protocol = port.get("protocol", "")  # "tcp" or "udp"

            # Service identification (populated by -sV)
            svc = port.find("service")
            service = product = version = extrainfo = ""
            if svc is not None:
                service   = svc.get("name", "")
                product   = svc.get("product", "")
                version   = svc.get("version", "")
                extrainfo = svc.get("extrainfo", "")

            # NSE script output — flatten all scripts for this port into one
            # pipe-delimited string (e.g. http-title, ssl-cert, smb-security-mode)
            script_parts = []
            for script in port.findall("script"):
                sid = script.get("id", "")
                out = script.get("output", "").strip().replace("\n", " ")
                if out:
                    script_parts.append(f"[{sid}] {out}")
            scripts = "  |  ".join(script_parts)

            rec = _base_record(ip, hostname, mac, os_name, scan_args)
            rec.update(
                port=port_id, protocol=protocol, state=state,
                service=service, product=product, version=version,
                extrainfo=extrainfo, scripts=scripts,
            )
            records.append(rec)

        if not open_found:
            # Host responded to the scan but all ports were closed/filtered
            records.append(_base_record(ip, hostname, mac, os_name, scan_args))

    return records


def _base_record(ip, hostname, mac, os_name, scan_args="") -> dict:
    """Return an empty port record pre-populated with host-level fields."""
    return {
        "ip": ip, "hostname": hostname, "mac": mac, "os": os_name,
        "scan_args": scan_args,
        "port": "", "protocol": "", "state": "up (no open ports)",
        "service": "", "product": "", "version": "", "extrainfo": "",
        "scripts": "",
        "location": "", "ew_title": "", "ew_status": "",
    }


# ---------------------------------------------------------------------------
# EyeWitness SQLite parser
# ---------------------------------------------------------------------------

def parse_eyewitness_db(db_path: Path) -> dict:
    """
    Parse an EyeWitness SQLite database and return a lookup dict.

    Returns:
        { (ip_str, port_str): {"title": str, "http_status": str}, ... }

    The EyeWitness schema has changed across versions; the parser handles
    both URL-style host fields (http://1.2.3.4:8080/) and raw IP fields,
    and accepts either 'http_status' or 'status_code' column names.
    """
    ew: dict = {}
    try:
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()

        # Verify the expected table exists before querying
        cur.execute("SELECT name FROM sqlite_master WHERE type='table'")
        tables = {r[0] for r in cur.fetchall()}
        if "hosts" not in tables:
            print(f"  WARNING: no 'hosts' table found in {db_path.name}")
            conn.close()
            return ew

        cur.execute("SELECT * FROM hosts")
        for row in cur.fetchall():
            row = dict(row)

            # 'host' column contains either a URL or a bare IP depending on EW version
            host_val = (row.get("host") or row.get("url") or "").strip()
            port_val = str(row.get("port") or "")

            ip = port = ""

            if "://" in host_val:
                # Parse http://ip:port/ or https://ip/ style entries
                m = re.match(r"https?://([^:/\[\]]+)(?::(\d+))?", host_val)
                if m:
                    ip = m.group(1)
                    if m.group(2):
                        port = m.group(2)
                    else:
                        # Infer default port from scheme when not explicit in URL
                        port = "443" if host_val.startswith("https") else "80"
                    # Prefer the dedicated port column if populated
                    port = port_val or port
            else:
                ip   = host_val
                port = port_val

            if not ip:
                continue

            title       = str(row.get("title") or "")
            http_status = str(
                row.get("http_status") or row.get("status_code") or ""
            )

            ew[(ip, port)] = {"title": title, "http_status": http_status}

        conn.close()
    except Exception as exc:
        print(f"  WARNING: EyeWitness parse error ({db_path.name}): {exc}")
    return ew


# ---------------------------------------------------------------------------
# Deduplication / merge
# ---------------------------------------------------------------------------

def deduplicate(records: list[dict]) -> list[dict]:
    """
    Merge records that share the same (ip, port, protocol) key.

    When two records describe the same port, the richer record (most non-empty
    fields) is used as the base and any fields it is missing are filled in from
    the other record. This ensures that information from multiple scan runs
    (e.g. a version scan and a default scan) is combined into a single row
    rather than dropped.
    """
    best: dict = {}
    for rec in records:
        key = (rec["ip"], rec["port"], rec["protocol"])
        if key not in best:
            best[key] = dict(rec)
        else:
            existing    = best[key]
            richness_new = sum(1 for v in rec.values() if v)
            richness_old = sum(1 for v in existing.values() if v)
            # Use the richer record as the base; fill gaps from the other
            base, other = (rec, existing) if richness_new >= richness_old else (existing, rec)
            merged = {k: base[k] if base[k] else other[k] for k in base}
            best[key] = merged
    return list(best.values())


def ip_sort_key(rec: dict) -> tuple:
    """
    Sort key that orders records numerically by IP octets then by port number.
    Falls back gracefully for IPv6 or malformed addresses.
    """
    parts  = rec["ip"].split(".")
    octets = []
    for p in parts:
        try:
            octets.append(int(p))
        except ValueError:
            octets.append(0)
    while len(octets) < 4:
        octets.append(0)
    port = int(rec["port"]) if str(rec["port"]).isdigit() else 0
    return (*octets, port)


# ---------------------------------------------------------------------------
# Scan statistics extraction
# ---------------------------------------------------------------------------

def extract_scan_stats(xml_path: Path) -> dict:
    """
    Extract scan-level statistics from a single nmap XML file.

    Returns a dict with the following keys (all IP-based values are sets so
    they deduplicate correctly when multiple scan files are merged):

      hosts_total           — total IPs nmap attempted (from <runstats>)
      hosts_up_ips          — set of IPs nmap marked as 'up'
      hosts_responsive_ips  — set of IPs with ≥1 open or closed port
                              (reliable 'reachable' indicator even with -Pn,
                               because a TCP RST/SYN-ACK proves the host exists;
                               all-filtered means we simply don't know)
      hosts_with_open_ports — set of IPs with ≥1 open port
      ports_open            — count of ports in state 'open'
      ports_open_filtered   — count of ports in state 'open|filtered'
      ports_filtered        — count of ports in state 'filtered'
      ports_closed          — count of ports in state 'closed'
      tcp_open              — open ports on TCP
      udp_open              — open ports on UDP
      unique_services       — set of service names identified (from -sV)
      pn_used               — True if -Pn was present in the scan arguments
                              (when True, hosts_up_ips == all targets, making
                               it an unreliable 'reachable' metric)
    """
    stats = {
        "hosts_total":           0,
        "hosts_up_ips":          set(),
        "hosts_responsive_ips":  set(),
        "hosts_with_open_ports": set(),
        "ports_open":            0,
        "ports_open_filtered":   0,
        "ports_filtered":        0,
        "ports_closed":          0,
        "tcp_open":              0,
        "udp_open":              0,
        "unique_services":       set(),
        "pn_used":               False,
    }

    try:
        root = ET.parse(xml_path).getroot()
    except ET.ParseError:
        return stats

    # Detect -Pn in the scan arguments stored in the XML root element.
    # When -Pn is active nmap skips host discovery and marks all targets as
    # 'up', so the hosts_up count from runstats equals hosts_total — useless
    # as a reachability metric.
    scan_args = root.get("args", "")
    if re.search(r"(?<!\w)-Pn(?!\w)", scan_args):
        stats["pn_used"] = True

    # Pull the total-hosts count from <runstats>; used for scope fallback only.
    runstats = root.find("runstats")
    if runstats is not None:
        hosts_el = runstats.find("hosts")
        if hosts_el is not None:
            stats["hosts_total"] = int(hosts_el.get("total", 0))

    # Walk every host nmap recorded as 'up' and tally port states and services.
    for host in root.findall("host"):
        status_el = host.find("status")
        if status_el is None or status_el.get("state") != "up":
            continue

        ip = ""
        for addr in host.findall("address"):
            if addr.get("addrtype") in ("ipv4", "ipv6"):
                ip = addr.get("addr", "")
                break

        if ip:
            stats["hosts_up_ips"].add(ip)

        ports_el = host.find("ports")
        if ports_el is None:
            continue

        for port in ports_el.findall("port"):
            state_el = port.find("state")
            state    = state_el.get("state", "") if state_el is not None else ""
            protocol = port.get("protocol", "")

            if state == "open":
                stats["ports_open"] += 1
                if protocol == "tcp":
                    stats["tcp_open"] += 1
                elif protocol == "udp":
                    stats["udp_open"] += 1
                if ip:
                    # open counts as both responsive and having an open port
                    stats["hosts_responsive_ips"].add(ip)
                    stats["hosts_with_open_ports"].add(ip)
                svc = port.find("service")
                if svc is not None:
                    name = svc.get("name", "")
                    if name:
                        stats["unique_services"].add(name)

            elif state == "open|filtered":
                stats["ports_open_filtered"] += 1
                if ip:
                    stats["hosts_with_open_ports"].add(ip)

            elif state == "filtered":
                stats["ports_filtered"] += 1

            elif state == "closed":
                stats["ports_closed"] += 1
                if ip:
                    # A closed port (TCP RST received) proves the host exists,
                    # so count it as responsive even though no port is open.
                    stats["hosts_responsive_ips"].add(ip)

    return stats


def merge_stats(stats_list: list[dict]) -> dict:
    """
    Combine statistics from multiple nmap scan files for the same location.

    IP-based fields use set unions so the same host is never counted twice
    across multiple scan files. Numeric port counts are summed (ports are
    per-scan observations, not per-host unique values).
    """
    merged = {
        "hosts_total":           0,
        "hosts_up_ips":          set(),
        "hosts_responsive_ips":  set(),
        "hosts_with_open_ports": set(),
        "ports_open":            0,
        "ports_open_filtered":   0,
        "ports_filtered":        0,
        "ports_closed":          0,
        "tcp_open":              0,
        "udp_open":              0,
        "unique_services":       set(),
        "pn_used":               False,
    }
    for s in stats_list:
        merged["hosts_total"]           += s["hosts_total"]
        merged["ports_open"]            += s["ports_open"]
        merged["ports_open_filtered"]   += s["ports_open_filtered"]
        merged["ports_filtered"]        += s["ports_filtered"]
        merged["ports_closed"]          += s["ports_closed"]
        merged["tcp_open"]              += s["tcp_open"]
        merged["udp_open"]              += s["udp_open"]
        # Set unions — deduplicates IPs and service names across files
        merged["hosts_up_ips"].update(s["hosts_up_ips"])
        merged["hosts_responsive_ips"].update(s["hosts_responsive_ips"])
        merged["hosts_with_open_ports"].update(s["hosts_with_open_ports"])
        merged["unique_services"].update(s["unique_services"])
        # Flag the location if any of its scan files used -Pn
        merged["pn_used"] = merged["pn_used"] or s["pn_used"]
    return merged


# ---------------------------------------------------------------------------
# Excel sheet builder
# ---------------------------------------------------------------------------

# Column definitions: (header label, column width in character units)
# EW_COLUMNS are appended only when EyeWitness data is present.
BASE_COLUMNS = [
    ("Source Location",   18),
    ("Target IP",         15),
    ("Hostname",          24),
    ("Port",               8),
    ("Protocol",          10),
    ("State",             16),
    ("Service",           14),
    ("Product / Banner",  26),
    ("Version",           16),
    ("Extra Info",        26),
    ("OS Detection",      26),
    ("NSE Script Output", 45),
    ("MAC Address",       18),
]

EW_COLUMNS = [
    ("Web Title",   32),
    ("HTTP Status", 12),
]


def write_sheet(ws, rows: list[dict], has_eyewitness: bool) -> None:
    """
    Write a complete formatted data table to the given worksheet.

    Applies header styling, alternating row shading, port-based colour
    coding, cell borders, and freezes the top row for easy scrolling.
    """
    columns = BASE_COLUMNS + (EW_COLUMNS if has_eyewitness else [])

    # ── Header row ──────────────────────────────────────────────────────────
    for col_idx, (label, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.border    = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"  # Keep header visible when scrolling

    # ── Data rows ───────────────────────────────────────────────────────────
    for row_idx, rec in enumerate(rows, 2):
        is_alt = row_idx % 2 == 0
        pfill  = port_fill(
            rec.get("port", ""), rec.get("protocol", ""), rec.get("state", "")
        )

        values = [
            rec.get("location",  ""),
            rec.get("ip",        ""),
            rec.get("hostname",  ""),
            rec.get("port",      ""),
            rec.get("protocol",  "").upper(),
            rec.get("state",     ""),
            rec.get("service",   ""),
            rec.get("product",   ""),
            rec.get("version",   ""),
            rec.get("extrainfo", ""),
            rec.get("os",        ""),
            rec.get("scripts",   ""),
            rec.get("mac",       ""),
        ]
        if has_eyewitness:
            values += [rec.get("ew_title", ""), rec.get("ew_status", "")]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border    = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            # Port colour takes priority over the alternating row shade
            if pfill:
                cell.fill = pfill
            elif is_alt:
                cell.fill = ALT_FILL


# ---------------------------------------------------------------------------
# Interactive file → location mapper
# ---------------------------------------------------------------------------

def prompt_mapping(files: list[Path], label: str) -> dict[str, list[Path]]:
    """
    Interactively prompt the user to assign each discovered file to a named
    scanning location. Multiple files can be assigned to the same location
    name — they will be merged and deduplicated during processing.
    Pressing Enter without a name skips the file.
    """
    if not files:
        return {}

    print(f"\n{'=' * 64}")
    print(f"  Found {len(files)} {label} file(s).")
    print("  Assign each to a scanning location name.")
    print("  Press Enter (blank) to skip a file.\n")

    mapping: dict[str, list[Path]] = defaultdict(list)
    for f in files:
        print(f"  File : {f}")
        loc = input("  Location (e.g. Azure, Mason, Lansing, Detroit): ").strip()
        if loc:
            mapping[loc].append(f)
        print()

    return dict(mapping)


# ---------------------------------------------------------------------------
# Statistics sheet
# ---------------------------------------------------------------------------

# Stat row definitions — the "IPs Targeted" row is prepended dynamically in
# write_stats_sheet so it can reference the scope count when one is provided.
#
# "Hosts w/ Definitive Response" counts IPs that produced at least one open or
# closed port — a TCP RST or SYN-ACK proves the host is reachable regardless of
# whether -Pn was used. This is the most reliable reachability metric available
# from nmap data; "Hosts Up (nmap)" is shown alongside it for reference.
_STAT_ROWS_BASE = [
    ("Hosts Up (nmap)",                  lambda s: len(s["hosts_up_ips"])),
    ("Hosts w/ Definitive Response",     lambda s: len(s["hosts_responsive_ips"])),
    ("Hosts w/ Open Ports",              lambda s: len(s["hosts_with_open_ports"])),
    ("Hosts w/ No Open Ports",           lambda s: max(0, len(s["hosts_responsive_ips"]) - len(s["hosts_with_open_ports"]))),
    ("",                                 lambda s: ""),   # blank separator row
    ("Total Open Ports",                 lambda s: s["ports_open"]),
    ("  ↳ TCP Open",                     lambda s: s["tcp_open"]),
    ("  ↳ UDP Open",                     lambda s: s["udp_open"]),
    ("Total Open|Filtered Ports",        lambda s: s["ports_open_filtered"]),
    ("Total Filtered Ports",             lambda s: s["ports_filtered"]),
    ("Total Closed Ports",               lambda s: s["ports_closed"]),
    ("",                                 lambda s: ""),   # blank separator row
    ("Unique Services Identified",       lambda s: len(s["unique_services"])),
]

# Fills used in the stats sheet header row and totals column
STATS_HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")  # dark blue — matches main sheets
STATS_TOTALS_FILL  = PatternFill("solid", fgColor="D6DCE4")  # light grey for the totals column
STATS_SECTION_FONT = Font(bold=True, size=11)


def write_stats_sheet(
    wb,
    stats_by_location: dict[str, dict],
    insert_index: int = 1,
    scope_count: int | None = None,
) -> None:
    """
    Insert a 'Statistics' sheet into the workbook at insert_index.

    Layout: rows are individual statistics; columns are scanning locations
    with a final 'Total' column that sums (or unions) across all locations.

        Statistic                       | Azure | Mason | ... | Total
        --------------------------------+-------+-------+-----+------
        IPs Targeted                    |  194  |  194  | ... |  194
        Hosts Up (nmap)                 |   14  |    4  | ... |   18
        Hosts w/ Definitive Response    |   12  |    4  | ... |   16
        Hosts w/ Open Ports             |    8  |    2  | ... |   10
        ...

    "Hosts Up (nmap)" = unique IPs nmap marked as up (unreliable with -Pn).
    "Hosts w/ Definitive Response" = IPs with ≥1 open or closed port;
      a TCP RST or SYN-ACK proves the host is reachable regardless of -Pn.

    scope_count: if provided (parsed from a scope.txt), this value is shown
    as "IPs Targeted" for every location column (same scope). If None, falls
    back to nmap runstats total with an explanatory footnote.

    A -Pn warning row is added to the sheet whenever any scan file in any
    location was run with -Pn, explaining why "Hosts Up" may equal scope size.
    """
    ws = wb.create_sheet(title="Statistics", index=insert_index)

    locations = sorted(stats_by_location.keys())
    num_locs  = len(locations)

    # Build the full row list, inserting the correct "IPs Targeted" row first.
    # When a scope file was provided the lambda always returns scope_count,
    # making the value identical across all location columns (same scope).
    if scope_count is not None:
        targeted_label = "IPs Targeted (from scope file)"
        targeted_fn    = lambda s, _sc=scope_count: _sc  # noqa: E731
    else:
        targeted_label = "IPs Targeted * (nmap runstats)"
        targeted_fn    = lambda s: s["hosts_total"]       # noqa: E731

    stat_rows = [(targeted_label, targeted_fn)] + _STAT_ROWS_BASE

    # ── Column headers ───────────────────────────────────────────────────────
    # Col 1 = "Statistic", then one col per location, then "Total"
    headers = ["Statistic"] + locations + ["Total"]
    col_widths = [32] + [16] * num_locs + [14]

    for col_idx, (label, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill      = STATS_HEADER_FILL
        cell.font      = HEADER_FONT
        cell.border    = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "B2"  # Freeze the statistic label column and header row

    # Pre-compute totals across all locations
    total_stats = merge_stats(list(stats_by_location.values()))

    # ── Data rows ────────────────────────────────────────────────────────────
    for row_idx, (label, fn) in enumerate(stat_rows, 2):
        is_separator = label == ""
        is_indent    = label.startswith("  ↳")

        # Statistic label cell
        label_cell = ws.cell(row=row_idx, column=1, value=label)
        label_cell.border    = BORDER
        label_cell.alignment = Alignment(vertical="center")
        if not is_separator and not is_indent:
            label_cell.font = STATS_SECTION_FONT
        else:
            label_cell.font = Font(size=11, italic=is_indent)

        # One value cell per location
        for col_idx, loc in enumerate(locations, 2):
            val  = fn(stats_by_location[loc]) if not is_separator else ""
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border    = BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font      = Font(size=11)
            # Shade alternating rows lightly for readability (skip separators)
            if not is_separator and row_idx % 2 == 0:
                cell.fill = ALT_FILL

        # Totals column — for the scope-file row, the total is just scope_count
        # (not a sum), since all locations share the same scope
        total_val  = fn(total_stats) if not is_separator else ""
        total_cell = ws.cell(row=row_idx, column=num_locs + 2, value=total_val)
        total_cell.fill      = STATS_TOTALS_FILL
        total_cell.border    = BORDER
        total_cell.alignment = Alignment(horizontal="center", vertical="center")
        total_cell.font      = Font(bold=True, size=11)

        ws.row_dimensions[row_idx].height = 18

    # ── Notes section ────────────────────────────────────────────────────────
    # Notes are appended below the data rows. Each note is a separate merged
    # row so they remain readable without fixed row heights.
    note_row = len(stat_rows) + 3
    note_fill = PatternFill("solid", fgColor="FFF9E6")  # pale yellow background
    note_font = Font(size=9, italic=True, color="555555")

    def _add_note(row, text):
        cell = ws.cell(row=row, column=1, value=text)
        cell.font      = note_font
        cell.fill      = note_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(
            start_row=row, start_column=1,
            end_row=row,   end_column=num_locs + 2,
        )
        ws.row_dimensions[row].height = 42
        return row + 1

    # Show runstats caveat only when no scope file was provided
    if scope_count is None:
        note_row = _add_note(
            note_row,
            "* IPs Targeted is sourced from nmap's <runstats> element and summed "
            "across all scan files assigned to each location. If multiple scan files "
            "cover overlapping target ranges, this count may be higher than the actual "
            "number of unique IPs. For a more accurate count, re-run and provide a scope file.",
        )

    # Show -Pn warning if any location's scans were run with -Pn
    any_pn = any(s["pn_used"] for s in stats_by_location.values())
    if any_pn:
        pn_locs = ", ".join(
            loc for loc in sorted(stats_by_location) if stats_by_location[loc]["pn_used"]
        )
        _add_note(
            note_row,
            f"⚠  -Pn detected in scan files for: {pn_locs}.  "
            "When -Pn is used nmap skips host discovery and marks every target as 'up', "
            "so 'Hosts Up (nmap)' will equal the total scope size — not a meaningful "
            "reachability indicator.  "
            "'Hosts w/ Definitive Response' (open or closed port observed) is the "
            "reliable metric in this case, as a TCP RST or SYN-ACK confirms the host exists.",
        )


# ---------------------------------------------------------------------------
# Legend sheet
# ---------------------------------------------------------------------------

def write_legend(wb) -> None:
    """
    Add a 'Legend' sheet to the workbook explaining the colour coding and
    listing all ports that trigger red/orange highlighting, split by protocol.
    """
    ws = wb.create_sheet(title="Legend")

    # Colour key entries: (label, description)
    entries = [
        ("Colour",           "Meaning"),
        ("Red cell",         "Sensitive / high-risk port — state: open (TCP or UDP specific)"),
        ("Orange cell",      "Sensitive / high-risk port — state: open|filtered"),
        ("Blue cell",        "Web / HTTP service port (any state)"),
        ("No highlight",     "Standard / low-risk port"),
        ("Alternating grey", "Even-numbered rows — no special significance"),
    ]

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 64

    fills = [HEADER_FILL, RED_FILL, ORANGE_FILL, BLUE_FILL,
             PatternFill("solid", fgColor="FFFFFF"), ALT_FILL]
    fonts = [HEADER_FONT] + [Font(size=11)] * 5

    for row_idx, ((label, desc), fill, font) in enumerate(
        zip(entries, fills, fonts), 1
    ):
        for col_idx, val in enumerate([label, desc], 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.fill      = fill
            c.font      = font
            c.border    = BORDER
            c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[row_idx].height = 22

    # Append the full list of flagged ports for each protocol
    r = len(entries) + 2

    ws.cell(row=r, column=1, value="Sensitive TCP ports:").font = Font(bold=True)
    r += 1
    c = ws.cell(row=r, column=1,
                value=", ".join(str(p) for p in sorted(SENSITIVE_PORTS_TCP)))
    c.alignment = Alignment(wrap_text=True)
    ws.merge_cells(f"A{r}:B{r}")
    ws.row_dimensions[r].height = 30

    r += 2
    ws.cell(row=r, column=1, value="Sensitive UDP ports:").font = Font(bold=True)
    r += 1
    c = ws.cell(row=r, column=1,
                value=", ".join(str(p) for p in sorted(SENSITIVE_PORTS_UDP)))
    c.alignment = Alignment(wrap_text=True)
    ws.merge_cells(f"A{r}:B{r}")
    ws.row_dimensions[r].height = 22


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> None:
    ap = argparse.ArgumentParser(
        description="Segmentation test report generator: nmap XML (+ EyeWitness) → Excel"
    )
    ap.add_argument(
        "directory", nargs="?", default=".",
        help="Root directory to search for input files (default: current dir)",
    )
    ap.add_argument(
        "-o", "--output", default=None,
        help=(
            "Output path. Accepts a bare filename ('report.xlsx') or a full "
            "absolute path. Bare filenames are written into the scanned "
            "directory. Default: segmentation_report.xlsx in the scanned directory."
        ),
    )
    args = ap.parse_args()

    root = Path(args.directory).resolve()
    if not root.exists():
        sys.exit(f"ERROR: {root} does not exist")

    # ── Discover input files ───────────────────────────────────────────────
    print(f"\nSearching for input files under: {root}")
    nmap_xmls = find_nmap_xmls(root)
    ew_dbs    = find_eyewitness_dbs(root)

    print(f"  nmap XML files  : {len(nmap_xmls)}")
    print(f"  EyeWitness DBs  : {len(ew_dbs)}")

    if not nmap_xmls:
        sys.exit("ERROR: No nmap XML files found under the given directory.")

    # ── Assign files to scanning locations (interactive) ───────────────────
    nmap_map = prompt_mapping(nmap_xmls, "nmap XML")
    ew_map   = prompt_mapping(ew_dbs, "EyeWitness DB") if ew_dbs else {}

    # ── Scope file (optional — for accurate IPs Targeted count) ───────────
    # The same scope applies to all locations, so one file is sufficient.
    scope_count = prompt_scope_file()

    if not nmap_map:
        sys.exit("No files were assigned to a location — nothing to do.")

    # ── Parse EyeWitness databases ─────────────────────────────────────────
    # Results are stored per location so they can be joined to the correct
    # nmap records later.
    ew_by_loc: dict[str, dict] = {}
    for loc, dbs in ew_map.items():
        combined: dict = {}
        for db in dbs:
            print(f"  [EyeWitness] Parsing {db.name}  →  {loc}")
            combined.update(parse_eyewitness_db(db))
        ew_by_loc[loc] = combined
    has_ew = bool(ew_by_loc)

    # ── Parse nmap, deduplicate per location, attach EyeWitness data ───────
    all_records: list[dict]        = []
    by_location: dict[str, list[dict]] = {}
    stats_by_location: dict[str, dict] = {}

    for loc in sorted(nmap_map):
        loc_recs: list[dict]   = []
        loc_stats: list[dict]  = []

        for xml_path in nmap_map[loc]:
            print(f"  [nmap] Parsing {xml_path.name}  →  {loc}")
            loc_recs.extend(parse_nmap_xml(xml_path))
            # Extract statistics separately so closed/filtered counts are
            # captured even though the main parser only keeps open ports
            loc_stats.append(extract_scan_stats(xml_path))

        # Merge duplicates that appear across multiple scan files for this location
        loc_recs = deduplicate(loc_recs)
        loc_recs.sort(key=ip_sort_key)

        # Join EyeWitness web title/status onto matching IP+port rows
        ew_loc = ew_by_loc.get(loc, {})
        for rec in loc_recs:
            rec["location"] = loc
            hit = ew_loc.get((rec["ip"], rec["port"]), {})
            rec["ew_title"]  = hit.get("title", "")
            rec["ew_status"] = hit.get("http_status", "")

        by_location[loc]      = loc_recs
        stats_by_location[loc] = merge_stats(loc_stats)
        all_records.extend(loc_recs)

    if not all_records:
        sys.exit("WARNING: No results to report (no open ports found).")

    # ── Build Excel workbook ───────────────────────────────────────────────
    wb = openpyxl.Workbook()

    # Summary sheet containing all locations combined (index 0)
    ws_all = wb.active
    ws_all.title = "All Results"
    write_sheet(ws_all, all_records, has_ew)

    # Statistics sheet inserted at index 1, right after All Results
    write_stats_sheet(wb, stats_by_location, insert_index=1, scope_count=scope_count)

    # One sheet per location, sorted alphabetically
    for loc in sorted(by_location):
        ws = wb.create_sheet(title=loc[:31])  # Excel enforces a 31-char sheet name limit
        write_sheet(ws, by_location[loc], has_ew)

    write_legend(wb)

    # ── Resolve output path ────────────────────────────────────────────────
    if args.output is None:
        # Default: write the report into the directory that was scanned
        out_path = root / "segmentation_report.xlsx"
    else:
        out_path = Path(args.output)
        if not out_path.is_absolute() and not out_path.parent.parts[1:]:
            # Bare filename given — place it in the scanned directory
            out_path = root / out_path

    wb.save(out_path)

    # ── Print summary to terminal ──────────────────────────────────────────
    print(f"\nReport saved → {out_path.resolve()}")
    print(f"\n{'Location':<22} {'Hosts':>7} {'Open Ports':>11}")
    print("─" * 42)
    for loc in sorted(by_location):
        recs  = by_location[loc]
        hosts = len({r["ip"] for r in recs})
        ports = len([r for r in recs if str(r["port"]).isdigit()])
        print(f"{loc:<22} {hosts:>7} {ports:>11}")
    print("─" * 42)
    total_hosts = len({r["ip"] for r in all_records})
    total_ports = len([r for r in all_records if str(r["port"]).isdigit()])
    print(f"{'TOTAL':<22} {total_hosts:>7} {total_ports:>11}")


if __name__ == "__main__":
    main()
